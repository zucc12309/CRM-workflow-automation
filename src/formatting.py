#!/usr/bin/env python3
"""
Nomura FI Client Engagement Pipeline — Excel Formatting Script v2
Creates professional multi-sheet Excel report with executive dashboard.
Usage: python formatting.py <clean_file_path> <json_metadata>
"""

import os
import sys
import json
import logging
import traceback
from datetime import datetime
from pathlib import Path

os.environ.setdefault("MPLCONFIGDIR", "/tmp/mplconfig")
Path(os.environ["MPLCONFIGDIR"]).mkdir(parents=True, exist_ok=True)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import (
    AxDataSource,
    NumData,
    NumDataSource,
    NumRef,
    NumVal,
    StrData,
    StrRef,
    StrVal,
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

# ─── Logging ─────────────────────────────────────────────────────────────────
LOG_PATH = "/Users/priyanshupatel/nomura_tmp/errors.log"
logging.basicConfig(
    filename=LOG_PATH, level=logging.ERROR,
    format="%(asctime)s [FORMATTING] %(levelname)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

# ─── Colour Palette ──────────────────────────────────────────────────────────
BASE_FONT = "Aptos"
TITLE_FONT = "Aptos Display"
CANVAS_BG = "F6F8FC"
PANEL_BG = "FFFFFF"
SLATE = "4B5563"
NAVY      = "1F3864"
NAVY_DK   = "142244"
BLUE      = "2E75B6"
BLUE_LT   = "EBF3FB"
BLUE_PAL  = "D6E4F0"
WHITE     = "FFFFFF"
RED       = "C00000"
RED_LT    = "FFCCCC"
RED_TXT   = "CC0000"
AMBER     = "8B6914"
AMBER_LT  = "FFF3CD"
AMBER_TXT = "856404"
GREEN_LT  = "D4EDDA"
GREEN_TXT = "155724"
GOLD      = "C9A84C"
TEAL      = "2196A6"
ORANGE    = "E65100"
PURPLE    = "6A1B9A"
GREY_BOR  = "CCCCCC"
GREY_LT   = "F7F9FC"
CARD_BG   = PANEL_BG
SEP_COLOR = "BDD7EE"

# ─── Layout Constants ────────────────────────────────────────────────────────
# Cols A–O (1–15): KPI cards + charts
# Col  P  (16):   visual separator
# Cols Q–T (17–20): mini-tables watchlist
TBL_S = 17   # Q
TBL_E = 20   # T

# ChartData sheet reference — set during dashboard build
_chart_ws   = [None]   # the hidden _ChartData worksheet
_chart_off  = [0]      # row counter on _ChartData sheet

FILTER_PANEL_ROW0 = 66


# ─── Style Helpers ───────────────────────────────────────────────────────────
def _fill(hex_c):
    return PatternFill(fill_type="solid", fgColor=hex_c)

def _font(size=10, bold=False, color="000000", italic=False):
    return Font(name=BASE_FONT, size=size, bold=bold, color=color, italic=italic)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin(color=GREY_BOR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _side(style="medium", color=BLUE):
    return Side(style=style, color=color)


def _mpl(color):
    text = str(color)
    return text if text.startswith("#") else f"#{text}"


def _write(ws, row, col, value, bg=None, fnt=None, aln=None, bdr=None, h=None, nfmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if bg:   c.fill         = _fill(bg)
    if fnt:  c.font         = fnt
    if aln:  c.alignment    = aln
    if bdr:  c.border       = bdr
    if nfmt: c.number_format = nfmt
    if h:    ws.row_dimensions[row].height = h
    return c


def _merge(ws, r1, c1, r2, c2, value=None, bg=None, fnt=None, aln=None, h=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return _write(ws, r1, c1, value, bg=bg, fnt=fnt, aln=aln, h=h)


# ─── Banner Row ──────────────────────────────────────────────────────────────
def _banner(ws, row, c1, c2, text, bg, fg, size=10, bold=True, height=20, italic=False):
    _merge(ws, row, c1, row, c2, text,
           bg=bg, fnt=_font(size, bold, fg, italic=italic),
           aln=_align(), h=height)


# ─── KPI Card  (3 cols wide × 6 rows tall) ───────────────────────────────────
def _kpi_card(ws, top, left, label, value, val_color=NAVY, sub=""):
    rc = left + 2
    bot = top + 5
    # Label stripe
    _merge(ws, top, left, top, rc, label.upper(),
           bg=NAVY_DK, fnt=_font(8.5, True, WHITE), aln=_align(), h=16)
    # Value block (3 rows)
    _merge(ws, top+1, left, top+3, rc, value,
           bg=CARD_BG, fnt=_font(24, True, val_color), aln=_align())
    for r in range(top+1, top+4):
        ws.row_dimensions[r].height = 18
    # Subtitle stripe
    _merge(ws, top+4, left, top+4, rc, sub,
           bg=BLUE_PAL, fnt=_font(7, False, SLATE, italic=True), aln=_align(), h=13)
    # Gap row
    ws.row_dimensions[top+5].height = 4
    # Border frame
    no = Side(style=None)
    med = _side("medium", BLUE)
    for r in range(top, bot+1):
        for col in range(left, rc+1):
            cell = ws.cell(row=r, column=col)
            cell.border = Border(
                left  = med if col == left else no,
                right = med if col == rc   else no,
                top   = med if r == top    else no,
                bottom= med if r == bot    else no,
            )


# ─── Section Label ───────────────────────────────────────────────────────────
def _sec(ws, row, c1, c2, text, bg=BLUE, fg=WHITE, h=18):
    _merge(ws, row, c1, row, c2, f"  {text}",
           bg=bg, fnt=_font(9.5, True, fg), aln=_align(h="left"), h=h)


# ─── Mini Table Builder ───────────────────────────────────────────────────────
def _mini_table(ws, start_row, headers, rows_data, title,
                title_bg=NAVY, flag_style=None, col_widths=None):
    """
    Draws a mini table starting at (start_row, TBL_S).
    flag_style: None=normal, 'red'=expired/critical, 'amber'=warning
    Returns next available row (after a 1-row gap).
    """
    n_cols = len(headers)
    end_col = TBL_S + n_cols - 1

    if col_widths:
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(TBL_S + i)].width = w

    # Title
    _merge(ws, start_row, TBL_S, start_row, end_col, title,
           bg=title_bg, fnt=_font(9, True, WHITE), aln=_align(), h=19)

    # Headers
    hdr_row = start_row + 1
    for i, h in enumerate(headers):
        _write(ws, hdr_row, TBL_S+i, h,
               bg=NAVY, fnt=_font(8, True, WHITE),
               aln=_align(), bdr=_thin(), h=17)

    # Empty fallback
    if not rows_data:
        _merge(ws, start_row+2, TBL_S, start_row+2, end_col,
               "✓  No items requiring attention",
               bg=GREEN_LT, fnt=_font(8, False, GREEN_TXT), aln=_align(), h=15)
        return start_row + 4

    # Data rows
    for r_idx, row_vals in enumerate(rows_data):
        rn = start_row + 2 + r_idx
        if flag_style == "red":
            bg_c, txt_c, bld = RED_LT, RED_TXT, True
        elif flag_style == "amber":
            bg_c, txt_c, bld = AMBER_LT, AMBER_TXT, True
        else:
            bg_c = WHITE if r_idx % 2 == 0 else GREY_LT
            txt_c, bld = SLATE, False

        for c_idx, val in enumerate(row_vals):
            halign = "left" if c_idx == 0 else "right"
            # Determine number format for numeric cells based on header
            hdr = headers[c_idx].lower() if c_idx < len(headers) else ""
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                if any(k in hdr for k in ("rev", "rpm", "mtg/$")):
                    nfmt = '"$"#,##0'
                elif "mtgs/$1k" in hdr or "mtgs/$" in hdr or "/" in hdr:
                    nfmt = '#,##0'
                else:
                    nfmt = '#,##0'
            else:
                nfmt = None
            _write(ws, rn, TBL_S+c_idx, val,
                   bg=bg_c, fnt=_font(8, bld, txt_c),
                   aln=_align(h=halign), bdr=_thin(), h=16, nfmt=nfmt)

    return start_row + 2 + len(rows_data) + 2  # +2 = gap row


def _chart_data(categories, values):
    """
    Write chart source data to the dedicated _ChartData sheet.
    Returns (cd_ws, base_row, n) where cd_ws is the chart data worksheet.
    """
    cd_ws = _chart_ws[0]
    base  = _chart_off[0] + 1          # 1-based row
    n     = min(len(categories), len(values))

    cd_ws.cell(row=base, column=1).value = "Category"
    cd_ws.cell(row=base, column=2).value = "Value"

    for i in range(n):
        cd_ws.cell(row=base + 1 + i, column=1).value = str(categories[i])
        cd_ws.cell(row=base + 1 + i, column=2).value = float(values[i])

    _chart_off[0] += n + 3             # gap of 2 blank rows between blocks
    return cd_ws, base, n


def _apply_chart_cache(chart, cd_ws, base, n, categories, values):
    """Embed inline data cache so Excel renders charts immediately on open."""
    if not chart.series or n <= 0:
        return

    cat_ref = str(Reference(cd_ws, min_col=1, min_row=base + 1, max_row=base + n))
    val_ref = str(Reference(cd_ws, min_col=2, min_row=base + 1, max_row=base + n))

    cat_cache = StrData(
        ptCount=n,
        pt=[StrVal(idx=i, v=str(categories[i])) for i in range(n)],
    )
    val_cache = NumData(
        ptCount=n,
        pt=[NumVal(idx=i, v=float(values[i])) for i in range(n)],
    )

    chart.series[0].cat = AxDataSource(strRef=StrRef(f=cat_ref, strCache=cat_cache))
    chart.series[0].val = NumDataSource(numRef=NumRef(f=val_ref, numCache=val_cache))


def _strip_chart_chrome(chart):
    """Remove gridlines from the VALUE axis only — never touch category axis."""
    # PieChart has no axes — skip
    if not hasattr(chart, "x_axis") or not hasattr(chart, "y_axis"):
        return
    # For horizontal (bar) charts: x_axis = value axis, y_axis = category axis (names)
    # For vertical (col) charts:   y_axis = value axis, x_axis = category axis (labels)
    # Only clear gridlines on the value axis so category labels stay visible.
    is_horiz = getattr(chart, "type", None) == "bar"
    value_axis = chart.x_axis if is_horiz else chart.y_axis
    if value_axis is not None:
        value_axis.majorGridlines = None
        value_axis.minorGridlines = None


def _bar(ws, anchor, categories, values, title, color=BLUE,
         width=13, height=9, horiz=False, show_labels=True, label_fmt='#,##0'):
    """Add a bar/column chart to ws. Data is stored on the _ChartData sheet."""
    cd_ws, base, n = _chart_data(categories, values)

    chart = BarChart()
    chart.type      = "bar" if horiz else "col"
    chart.grouping  = "clustered"
    chart.title     = title
    chart.style     = 10
    chart.width     = width
    chart.height    = height
    chart.legend    = None
    chart.gapWidth  = 80 if horiz else 60

    if horiz:
        # x_axis = value axis (numbers), y_axis = category axis (names)
        chart.x_axis.numFmt     = label_fmt
        chart.x_axis.title      = None
        chart.x_axis.delete     = False
        chart.y_axis.delete     = False
        chart.y_axis.tickLblPos = "low"
    else:
        # y_axis = value axis (numbers), x_axis = category axis (labels)
        chart.y_axis.numFmt = label_fmt
        chart.y_axis.title  = None
        chart.y_axis.delete = False
        chart.x_axis.delete = False

    data_ref = Reference(cd_ws, min_col=2, min_row=base + 1, max_row=base + n)
    cats_ref = Reference(cd_ws, min_col=1, min_row=base + 1, max_row=base + n)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)

    if show_labels:
        chart.dLbls               = DataLabelList()
        chart.dLbls.showVal       = True
        chart.dLbls.numFmt        = label_fmt
        chart.dLbls.showLegendKey = False
        chart.dLbls.showSerName   = False
        chart.dLbls.showCatName   = False

    chart.series[0].graphicalProperties.solidFill      = color
    chart.series[0].graphicalProperties.line.solidFill = color

    _apply_chart_cache(chart, cd_ws, base, n, categories, values)
    _strip_chart_chrome(chart)
    ws.add_chart(chart, anchor)


def _pie(ws, anchor, categories, values, title, colors=None, width=10, height=9):
    """Add a pie chart to ws. Data is stored on the _ChartData sheet."""
    cd_ws, base, n = _chart_data(categories, values)

    pie            = PieChart()
    pie.title      = title
    pie.style      = 10
    pie.width      = width
    pie.height     = height
    pie.varyColors = True
    pie.legend     = None

    data_ref = Reference(cd_ws, min_col=2, min_row=base + 1, max_row=base + n)
    cats_ref = Reference(cd_ws, min_col=1, min_row=base + 1, max_row=base + n)
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(cats_ref)

    pie.dLbls               = DataLabelList()
    pie.dLbls.showPercent   = True
    pie.dLbls.showCatName   = True
    pie.dLbls.showSerName   = False
    pie.dLbls.showLegendKey = False
    pie.dLbls.showVal       = False

    palette = colors or [GOLD, BLUE, TEAL, NAVY, ORANGE, PURPLE]
    for idx, c in enumerate(palette[:n]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = c
        pie.series[0].dPt.append(pt)

    _apply_chart_cache(pie, cd_ws, base, n, categories, values)
    _strip_chart_chrome(pie)
    ws.add_chart(pie, anchor)


def _filter_list(values):
    cleaned = []
    for val in values:
        text = str(val).strip()
        if not text or text.lower() == "nan":
            text = "Unknown"
        if text not in cleaned:
            cleaned.append(text)
    return ["All"] + sorted(cleaned, key=str.lower)


def _xl_quote(text):
    return str(text).replace('"', '""')


def _build_filter_lists_sheet(ws, df):
    ws.sheet_state = "hidden"

    fields = [
        ("Region", "REGION"),
        ("Analyst", "analyst_nm"),
        ("Tier", "client_tier"),
        ("MiFID", "MiFID_Status"),
        ("Flag", "Flag"),
    ]

    ranges = {}
    for col_idx, (label, field) in enumerate(fields, start=1):
        values = _filter_list(df[field].dropna().tolist()) if field in df.columns else ["All"]
        ws.cell(row=1, column=col_idx).value = label
        ws.cell(row=1, column=col_idx).font = _font(9, True, WHITE)
        ws.cell(row=1, column=col_idx).fill = _fill(NAVY)
        ws.cell(row=1, column=col_idx).alignment = _align()
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(label) + 2)

        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx).value = value

        ranges[field] = f"Filter_Lists!${get_column_letter(col_idx)}$2:${get_column_letter(col_idx)}${len(values)+1}"

    return ranges


def _criteria_formula(cell_ref):
    return f'IF({cell_ref}="All","*",{cell_ref})'


def _sumifs_formula(sum_range, criteria_pairs):
    parts = [f"{rng},{crit}" for rng, crit in criteria_pairs]
    return f'SUMIFS({sum_range},' + ",".join(parts) + ')'


def _countifs_formula(criteria_pairs):
    parts = [f"{rng},{crit}" for rng, crit in criteria_pairs]
    return f'COUNTIFS(' + ",".join(parts) + ')'


def _build_interactive_filters(ws, df, processing_date, total_rows_clean, list_ranges, row0=1, configure_layout=True):
    if configure_layout:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90

        headers = list(df.columns)
        col_map = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}
        # Layout
        for col in range(1, 21):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 4
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 14
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 14
        ws.column_dimensions["L"].width = 14
        ws.column_dimensions["M"].width = 14
        ws.column_dimensions["N"].width = 14
        ws.column_dimensions["O"].width = 14
    else:
        headers = list(df.columns)
        col_map = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}

    last_row = len(df) + 2
    last_col = get_column_letter(len(headers))

    def rr(n):
        return row0 + n - 1

    _banner(ws, rr(1), 1, 20,
            f"INTERACTIVE FILTERS  ·  {processing_date}",
            NAVY_DK, WHITE, size=15, bold=True, height=38)
    _banner(ws, rr(2), 1, 20,
            "Use the dropdowns below like Power BI slicers. Select All to reset the view.",
            BLUE_LT, SLATE, size=9.5, italic=True, bold=False, height=22)

    # Filter panel
    _sec(ws, rr(4), 1, 2, "FILTER CONTROLS", bg=NAVY_DK)
    controls = [
        ("Region", f"B{rr(5)}", "REGION"),
        ("Analyst", f"B{rr(6)}", "analyst_nm"),
        ("Tier", f"B{rr(7)}", "client_tier"),
        ("MiFID", f"B{rr(8)}", "MiFID_Status"),
        ("Flag", f"B{rr(9)}", "Flag"),
    ]

    for label, cell_ref, field in controls:
        row = ws[cell_ref].row
        _write(ws, row, 1, label, bg=WHITE, fnt=_font(9.5, True, NAVY),
               aln=_align(h="left"), bdr=_thin(), h=20)
        _write(ws, row, 2, "All", bg=BLUE_LT, fnt=_font(9.5, True, NAVY),
               aln=_align(), bdr=_thin(), h=20)
        rng = list_ranges.get(field)
        if rng:
            dv = DataValidation(type="list", formula1=f"={rng}", allow_blank=False)
            dv.promptTitle = f"Choose {label}"
            dv.prompt = "Select a value or leave All to show everything."
            dv.errorTitle = "Invalid selection"
            dv.error = "Use the dropdown values only."
            ws.add_data_validation(dv)
            dv.add(ws[cell_ref])

    _merge(ws, rr(11), 1, rr(11), 2,
           "Power BI-style slicing with live formulas and a filtered table below.",
           bg=BLUE_PAL, fnt=_font(8.5, False, SLATE, italic=True), aln=_align(), h=18)

    # Summary cards
    region_cell = f"$B${rr(5)}"
    analyst_cell = f"$B${rr(6)}"
    tier_cell = f"$B${rr(7)}"
    mifid_cell = f"$B${rr(8)}"
    flag_cell = f"$B${rr(9)}"

    count_pairs = [
        (f"Data!${col_map['REGION']}$3:${col_map['REGION']}${last_row}", _criteria_formula(region_cell)),
        (f"Data!${col_map['analyst_nm']}$3:${col_map['analyst_nm']}${last_row}", _criteria_formula(analyst_cell)),
        (f"Data!${col_map['client_tier']}$3:${col_map['client_tier']}${last_row}", _criteria_formula(tier_cell)),
        (f"Data!${col_map['MiFID_Status']}$3:${col_map['MiFID_Status']}${last_row}", _criteria_formula(mifid_cell)),
        (f"Data!${col_map['Flag']}$3:${col_map['Flag']}${last_row}", _criteria_formula(flag_cell)),
    ]
    rev_sum = _sumifs_formula(
        f"Data!${col_map['revenue_usd']}$3:${col_map['revenue_usd']}${last_row}",
        count_pairs,
    )
    mtg_sum = _sumifs_formula(
        f"Data!${col_map['meetings_cnt']}$3:${col_map['meetings_cnt']}${last_row}",
        count_pairs,
    )

    _kpi_card(ws, rr(4), 4, "Matching Clients", f"={_countifs_formula(count_pairs)}", NAVY, "Rows matching filters")
    _kpi_card(ws, rr(4), 7, "Matching Revenue", f'=TEXT({rev_sum},"$#,##0")', GREEN_TXT, "Filtered portfolio revenue")
    _kpi_card(ws, rr(4), 10, "Matching Meetings", f"={mtg_sum}", BLUE, "Filtered client meetings")
    _kpi_card(ws, rr(4), 13, "Avg Rev / Meeting", f'=IF({mtg_sum}=0,"$0",TEXT({rev_sum}/{mtg_sum},"$#,##0"))', AMBER_TXT, "Commercial efficiency")

    # Selected filters strip
    _merge(ws, rr(10), 4, rr(10), 15,
           f'="Region: " & $B${rr(5)} & "  •  Analyst: " & $B${rr(6)} & "  •  Tier: " & $B${rr(7)} & "  •  MiFID: " & $B${rr(8)} & "  •  Flag: " & $B${rr(9)}',
           bg=WHITE, fnt=_font(8.5, False, SLATE, italic=True), aln=_align(), h=18)

    # Live filtered table
    for idx, header in enumerate(headers, start=1):
        _write(ws, rr(13), idx, header, bg=NAVY, fnt=_font(9, True, WHITE), aln=_align(), bdr=_thin(), h=19)

    # Use (cell="All")+(range=cell)>0 — always produces an array.
    # IF(cell="All", 1, ...) returns scalar 1 → FILTER(range, 1) = "No matching records".
    r_rng  = f'Data!${col_map["REGION"]}$3:${col_map["REGION"]}${last_row}'
    an_rng = f'Data!${col_map["analyst_nm"]}$3:${col_map["analyst_nm"]}${last_row}'
    ti_rng = f'Data!${col_map["client_tier"]}$3:${col_map["client_tier"]}${last_row}'
    mf_rng = f'Data!${col_map["MiFID_Status"]}$3:${col_map["MiFID_Status"]}${last_row}'
    fl_rng = f'Data!${col_map["Flag"]}$3:${col_map["Flag"]}${last_row}'

    mask = (
        f'(({region_cell}="All")+({r_rng}={region_cell})>0)'
        f'*(({analyst_cell}="All")+({an_rng}={analyst_cell})>0)'
        f'*(({tier_cell}="All")+({ti_rng}={tier_cell})>0)'
        f'*(({mifid_cell}="All")+({mf_rng}={mifid_cell})>0)'
        f'*(({flag_cell}="All")+({fl_rng}={flag_cell})>0)'
    )
    spill_range = f"Data!$A$3:${last_col}${last_row}"
    ws[f"A{rr(14)}"] = f'=IFERROR(FILTER({spill_range},{mask}),"No matching records")'
    if row0 == 1:
        ws.freeze_panes = f"A{rr(14)}"


# ─── KPI Card v2 (wider, bolder, icon row) ───────────────────────────────────
def _kpi_card2(ws, top, left, label, value, val_color, sub, icon=""):
    """5-col wide × 6-row tall KPI card with icon + label stripe."""
    rc = left + 4   # 5 cols wide
    bot = top + 5

    # Top stripe: icon + label
    _merge(ws, top, left, top, rc,
           (icon + "  " if icon else "") + label.upper(),
           bg=NAVY_DK, fnt=_font(9, True, GOLD if icon else WHITE),
           aln=_align(h="left"), h=17)
    # Value rows (3 tall)
    _merge(ws, top+1, left, top+3, rc, value,
           bg=PANEL_BG, fnt=_font(28, True, val_color), aln=_align())
    for r in range(top+1, top+4):
        ws.row_dimensions[r].height = 20
    # Subtitle stripe
    _merge(ws, top+4, left, top+4, rc, sub,
           bg=BLUE_LT, fnt=_font(7.5, False, SLATE, italic=True),
           aln=_align(), h=13)
    # Gap row
    ws.row_dimensions[top+5].height = 5

    # Coloured left-edge accent + outer border
    no = Side(style=None)
    acc  = Side(style="thick", color=val_color if val_color != NAVY else BLUE)
    thin = Side(style="thin",  color=GREY_BOR)
    med  = Side(style="medium", color=GREY_BOR)
    for r in range(top, bot+1):
        for col in range(left, rc+1):
            cell = ws.cell(row=r, column=col)
            cell.border = Border(
                left   = acc  if col == left else thin,
                right  = med  if col == rc   else thin,
                top    = med  if r == top    else no,
                bottom = med  if r == bot    else no,
            )


# ─── Dashboard Sheet ─────────────────────────────────────────────────────────
def _build_dashboard(ws, df, processing_date, total_rows_clean, list_ranges):
    _chart_off[0] = 0   # reset row counter on _ChartData sheet

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85

    # ── Pre-compute stats ────────────────────────────────────────────────────
    last_row    = len(df) + 2
    total_rev   = df["revenue_usd"].sum()
    total_mtg   = int(df["meetings_cnt"].sum())
    avg_rpm     = round(total_rev / total_mtg, 2) if total_mtg > 0 else 0
    expired_cnt = int((df["MiFID_Status"] == "EXPIRED").sum())
    flagged_cnt = int((df["Flag"] == "Review").sum())

    # ── Background canvas ────────────────────────────────────────────────────
    for r in range(1, 90):
        for c in range(1, TBL_E + 1):
            ws.cell(row=r, column=c).fill = _fill("EEF2FA")

    # ── Column widths ────────────────────────────────────────────────────────
    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 8.8
    ws.column_dimensions["P"].width = 2.5
    ws.column_dimensions["Q"].width = 16
    ws.column_dimensions["R"].width = 20
    ws.column_dimensions["S"].width = 13
    ws.column_dimensions["T"].width = 13
    tbl_widths = [22, 14, 11, 11]
    for i, w in enumerate(tbl_widths):
        ws.column_dimensions[get_column_letter(TBL_S + i)].width = w

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 1-3: HEADER BLOCK
    # ════════════════════════════════════════════════════════════════════════
    _merge(ws, 1, 1, 1, TBL_E,
           "  NOMURA  \u00b7  FIXED INCOME CLIENT ENGAGEMENT DASHBOARD",
           bg=NAVY_DK, fnt=_font(20, True, WHITE), aln=_align(h="left"), h=52)

    _merge(ws, 2, 1, 2, TBL_E,
           f"  {processing_date}   |   {total_rows_clean} Active Clients   "
           f"|   Portfolio Revenue: ${total_rev:,.0f}   "
           f"|   MiFID Expired: {expired_cnt}   |   Clients Flagged: {flagged_cnt}",
           bg=BLUE, fnt=_font(10, False, WHITE), aln=_align(h="left"), h=22)

    # Gold accent line
    _merge(ws, 3, 1, 3, TBL_E, None, bg=GOLD, h=4)

    # Vertical separator col P
    for r in range(1, 87):
        ws.cell(row=r, column=16).fill = _fill("C8D8EE")

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 18-31: POWER BI FILTER PANEL (cols Q-T) — moved below KPI rows
    # ════════════════════════════════════════════════════════════════════════
    _merge(ws, 18, TBL_S, 18, TBL_E,
           "  FILTER PANEL",
           bg="0D1F3C", fnt=_font(10, True, GOLD), aln=_align(h="left"), h=23)

    _merge(ws, 19, TBL_S, 19, TBL_E,
           "  Use dropdowns to slice the live KPI metrics below.",
           bg=BLUE_LT, fnt=_font(8, False, SLATE, italic=True), aln=_align(h="left"), h=17)

    FILTER_DEFS = [
        ("Region",  "REGION",       20, "14365D"),
        ("Analyst", "analyst_nm",   21, "1C4E8A"),
        ("Tier",    "client_tier",  22, "1C6E6E"),
        ("MiFID",   "MiFID_Status", 23, "8B1A1A"),
        ("Flag",    "Flag",         24, "7B5900"),
    ]

    filter_col_letter = get_column_letter(TBL_S + 1)
    filter_cell_refs  = {}

    for label, field, row, stripe_color in FILTER_DEFS:
        _write(ws, row, TBL_S, label,
               bg=stripe_color, fnt=_font(9, True, WHITE),
               aln=_align(h="left"), bdr=_thin(stripe_color), h=22)
        _write(ws, row, TBL_S + 1, "All",
               bg=WHITE, fnt=_font(10, True, BLUE),
               aln=_align(), bdr=_thin(BLUE), h=22)
        ws.merge_cells(start_row=row, start_column=TBL_S + 2,
                       end_row=row, end_column=TBL_E)
        ws.cell(row=row, column=TBL_S + 2).fill = _fill(BLUE_LT)

        rng = list_ranges.get(field)
        if rng:
            dv = DataValidation(type="list", formula1=f"={rng}", allow_blank=False)
            dv.promptTitle = f"Filter by {label}"
            dv.prompt = "Select a value or choose All to clear this filter."
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=row, column=TBL_S + 1))

        filter_cell_refs[field] = f"${filter_col_letter}${row}"

    r_reg  = filter_cell_refs["REGION"]
    r_an   = filter_cell_refs["analyst_nm"]
    r_tier = filter_cell_refs["client_tier"]
    r_mif  = filter_cell_refs["MiFID_Status"]
    r_flg  = filter_cell_refs["Flag"]

    def _all_or(cell, col):
        """Returns 1 for each row matching the filter, or 1 for all rows if 'All'."""
        return f'(({cell}="All")+({col}={cell})>0)'

    # build column letter map from df column order
    headers = list(df.columns)
    col_map = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}

    D = "Data"  # sheet name shorthand
    r1 = 3      # first data row on Data sheet

    def _rng(col_name):
        return f"{D}!${col_map[col_name]}${r1}:${col_map[col_name]}${last_row}"

    reg_match  = _all_or(r_reg,  f"{D}!${col_map['REGION']}$3:${col_map['REGION']}${last_row}")
    an_match   = _all_or(r_an,   f"{D}!${col_map['analyst_nm']}$3:${col_map['analyst_nm']}${last_row}")
    tier_match = _all_or(r_tier, f"{D}!${col_map['client_tier']}$3:${col_map['client_tier']}${last_row}")
    mif_match  = _all_or(r_mif,  f"{D}!${col_map['MiFID_Status']}$3:${col_map['MiFID_Status']}${last_row}")
    flg_match  = _all_or(r_flg,  f"{D}!${col_map['Flag']}$3:${col_map['Flag']}${last_row}")

    mask = f"{reg_match}*{an_match}*{tier_match}*{mif_match}*{flg_match}"

    rev_col = f"{D}!${col_map['revenue_usd']}$3:${col_map['revenue_usd']}${last_row}"
    mtg_col = f"{D}!${col_map['meetings_cnt']}$3:${col_map['meetings_cnt']}${last_row}"

    cnt_f = f'=SUMPRODUCT({mask})'
    rev_f = f'=SUMPRODUCT({mask},{rev_col})'           # raw number — formatted via number_format
    mtg_f = f'=SUMPRODUCT({mask},{mtg_col})'
    rpm_f = f'=IFERROR(SUMPRODUCT({mask},{rev_col})/SUMPRODUCT({mask},{mtg_col}),0)'
    flg_f = f'=SUMPRODUCT({mask}*({D}!${col_map["Flag"]}$3:${col_map["Flag"]}${last_row}="Review"))'

    # Active filter summary strip (row 25)
    sel_f = (f'="Region: "&{r_reg}&"  |  Analyst: "&{r_an}'
             f'&"  |  Tier: "&{r_tier}&"  |  MiFID: "&{r_mif}&"  |  Flag: "&{r_flg}')
    _merge(ws, 25, TBL_S, 25, TBL_E, sel_f,
           bg="DCE8F8", fnt=_font(7.5, False, NAVY, italic=True), aln=_align(), h=16)

    # ── Live KPI summary driven by filter cells ──────────────────────────────
    _merge(ws, 26, TBL_S, 26, TBL_E,
           "  FILTERED KPIs  \u2014  live, updates with filters above",
           bg=TEAL, fnt=_font(8.5, True, WHITE), aln=_align(h="left"), h=20)

    live_kpis = [
        (27, "Matching Clients",   cnt_f, NAVY),
        (28, "Filtered Revenue",   rev_f, GREEN_TXT),
        (29, "Total Meetings",     mtg_f, BLUE),
        (30, "Avg Rev / Meeting",  rpm_f, AMBER_TXT),
        (31, "Flagged Clients",    flg_f, RED_TXT),
    ]
    _USD_FMT  = '"$"#,##0'          # locale-independent: $ hardcoded, comma = thousands sep
    _USD_FMT2 = '"$"#,##0.00'
    _kpi_nfmt = {28: _USD_FMT, 30: _USD_FMT}

    for row_n, lbl, formula, val_color in live_kpis:
        _write(ws, row_n, TBL_S, lbl,
               bg="F8FAFF", fnt=_font(8.5, False, SLATE),
               aln=_align(h="left"), bdr=_thin(), h=20)
        _merge(ws, row_n, TBL_S + 1, row_n, TBL_E, formula,
               bg=WHITE, fnt=_font(10, True, val_color),
               aln=_align(), h=20)
        if row_n in _kpi_nfmt:
            ws.cell(row=row_n, column=TBL_S + 1).number_format = _kpi_nfmt[row_n]
        for cc in range(TBL_S + 1, TBL_E + 1):
            ws.cell(row=row_n, column=cc).border = _thin()

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 4-9: KPI CARDS — 2 rows of 3 cards (5 cols wide each)
    # ════════════════════════════════════════════════════════════════════════
    def _kpi3(top, left, label, value, color, sub, tag):
        """5-col x 6-row premium KPI card."""
        rc  = left + 4
        bot = top + 5
        _merge(ws, top, left, top, rc,
               f"  {tag}  \u00b7  {label.upper()}",
               bg=NAVY_DK, fnt=_font(9, True, GOLD), aln=_align(h="left"), h=18)
        _merge(ws, top + 1, left, top + 3, rc, value,
               bg=WHITE, fnt=_font(26, True, color), aln=_align())
        for r in range(top + 1, top + 4):
            ws.row_dimensions[r].height = 20
        _merge(ws, top + 4, left, top + 4, rc, sub,
               bg=BLUE_LT, fnt=_font(7.5, False, SLATE, italic=True), aln=_align(), h=13)
        ws.row_dimensions[top + 5].height = 5
        no  = Side(style=None)
        acc = Side(style="thick", color=color)
        med = Side(style="medium", color=GREY_BOR)
        thn = Side(style="thin",  color=GREY_BOR)
        for r in range(top, bot + 1):
            for col in range(left, rc + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = Border(
                    left   = acc if col == left else thn,
                    right  = med if col == rc   else thn,
                    top    = med if r == top    else no,
                    bottom = med if r == bot    else no,
                )

    # Row 1 of KPI cards (rows 4-9)
    _kpi3(4,  1, "Total Clients",   str(total_rows_clean), NAVY,      "Active client records",     "CLIENTS")
    _kpi3(4,  6, "Total Revenue",   f"${total_rev:,.0f}",  GREEN_TXT, "USD across full portfolio", "REVENUE")
    _kpi3(4, 11, "Avg Rev / Mtg",   f"${avg_rpm:,.0f}",    BLUE,      "Portfolio efficiency",      "EFFICIENCY")

    # Row 2 of KPI cards (rows 10-15)
    _kpi3(10,  1, "MiFID Expired",  str(expired_cnt),      RED_TXT,   "Immediate compliance action",  "COMPLIANCE")
    _kpi3(10,  6, "Clients Flagged",str(flagged_cnt),      AMBER_TXT, "High meetings, low revenue",   "REVIEW")
    _kpi3(10, 11, "Total Meetings", f"{total_mtg:,}",      TEAL,      "Cumulative client meetings",   "MEETINGS")

    # Row 16: additional metrics summary strip (Issue 5)
    avg_eng = round(float(df["Engagement_Score"].mean()), 1) if "Engagement_Score" in df.columns else 0
    unknown_mifid = int((df["MiFID_Status"] == "Unknown").sum())
    total_eng = df["Engagement_Score"].sum() if "Engagement_Score" in df.columns else 1
    rev_per_eng = round(total_rev / total_eng, 0) if total_eng > 0 else 0

    _merge(ws, 16, 1, 16, 15,
           f"  Avg Engagement Score: {avg_eng}  \u00b7  MiFID Unknown: {unknown_mifid} clients  \u00b7  Rev / Engagement Point: ${rev_per_eng:,.0f}  \u00b7  Pending MiFID Review: {unknown_mifid}",
           bg="1A2E52", fnt=_font(9, False, "B8CCE8"), aln=_align(h="left"), h=18)

    # ════════════════════════════════════════════════════════════════════════
    # ROW 17: Section dividers
    # ════════════════════════════════════════════════════════════════════════
    _merge(ws, 17, 1, 17, 15,
           "  PORTFOLIO ANALYTICS  \u2014  Revenue  |  Analyst Performance  |  Client ROI  |  Risk",
           bg=NAVY, fnt=_font(10, True, WHITE), aln=_align(h="left"), h=22)
    # KEY WATCHLIST header lives at row 17 col Q-T — but the mini tables
    # start at row 32 (after the filter panel). Write a label here to match:
    _merge(ws, 17, TBL_S, 17, TBL_E,
           "  FILTER  &  KPI PANEL",
           bg="0D1F3C", fnt=_font(10, True, GOLD), aln=_align(h="left"), h=22)

    # ════════════════════════════════════════════════════════════════════════
    # CHARTS — static numeric values for guaranteed rendering
    # ════════════════════════════════════════════════════════════════════════

    # ── Chart sizing ─────────────────────────────────────────────────────────
    # Each col is 8.8 char wide. At Calibri 11pt: 1 char ≈ 0.182 cm.
    # Cols A–H (8 × 8.8): 12.8 cm left zone
    # Cols I–P (7 × 8.8 + 2.5): 11.6 cm right zone (must not overflow into Q)
    # Full width A–O (15 × 8.8 − col P gap): 22.5 cm
    CH_W_LEFT  = 12.5   # cm — left chart (A anchor)
    CH_W_RIGHT = 11.2   # cm — right chart (I anchor) — tight to stay within col P
    CH_H_PAIR  =  9.5   # cm — row height for each pair section
    CH_W_FULL  = 22.5   # cm — chart 5 full width A–O
    CH_H_FULL  =  9.0   # cm — chart 5 height

    # CHART 1: Revenue by Region (vertical bar) — A18
    reg = (df[df["REGION"] != "UNKNOWN"].groupby("REGION")["revenue_usd"].sum()
             .reset_index().sort_values("revenue_usd", ascending=False))
    _bar(ws, "A18",
         reg["REGION"].tolist(),
         [float(v) for v in reg["revenue_usd"]],
         "Revenue by Region (USD)",
         color=BLUE, width=CH_W_LEFT, height=CH_H_PAIR,
         show_labels=True, label_fmt='$#,##0')

    # CHART 2: Top Analysts by Revenue (horizontal bar) — I18
    an = (df[~df["analyst_nm"].isin(["System", "Unknown Analyst"])]
            .groupby("analyst_nm")["revenue_usd"].sum()
            .reset_index().sort_values("revenue_usd", ascending=True).tail(8))
    _bar(ws, "I18",
         an["analyst_nm"].tolist(),
         [float(v) for v in an["revenue_usd"]],
         "Top Analysts \u2014 Revenue (USD)",
         color=NAVY, width=CH_W_RIGHT, height=CH_H_PAIR, horiz=True,
         show_labels=True, label_fmt='$#,##0')

    # Section label before row 2 of charts
    _merge(ws, 38, 1, 38, 15,
           "  CLIENT RISK & EFFICIENCY  \u2014  Low ROI  |  Engagement vs Revenue  |  Tier Mix",
           bg=TEAL, fnt=_font(10, True, WHITE), aln=_align(h="left"), h=20)

    # CHART 3: Tier Pie — A39  (exclude Unspecified)
    tier_df = (df[df["client_tier"] != "Unspecified"]
                 .groupby("client_tier")["client_name"].count()
                 .reset_index().rename(columns={"client_name": "Count"}))
    _pie(ws, "A39",
         tier_df["client_tier"].tolist(),
         [float(v) for v in tier_df["Count"]],
         "Client Distribution by Tier",
         colors=[GOLD, BLUE, TEAL, NAVY, ORANGE, RED],
         width=CH_W_LEFT, height=CH_H_PAIR)

    # CHART 4: Low ROI — I39
    # Only include clients with actual revenue > 0 so zero-rev clients don't dominate
    df_roi = df[df["revenue_usd"] > 0].copy()
    df_roi["effort_per_1k"] = (
        df_roi["meetings_cnt"] / (df_roi["revenue_usd"] / 1000.0)
    ).round(1)
    waste = df_roi.nlargest(8, "effort_per_1k")
    _bar(ws, "I39",
         [n[:16] for n in waste["client_name"]],
         [float(v) for v in waste["effort_per_1k"]],
         "Low ROI: Meetings per $1K Revenue (excl. zero-rev)",
         color=RED, width=CH_W_RIGHT, height=CH_H_PAIR, horiz=True,
         show_labels=True, label_fmt='0.0" mtgs"')

    # Section label before analyst efficiency
    _merge(ws, 58, 1, 58, 15,
           "  ANALYST EFFICIENCY RANKING  \u2014  Revenue per Meeting  |  Sorted: Worst to Best",
           bg=PURPLE, fnt=_font(10, True, WHITE), aln=_align(h="left"), h=20)

    # CHART 5: Analyst Efficiency worst→best (full width) — A59
    eff = (df[~df["analyst_nm"].isin(["System", "Unknown Analyst"])]
             .groupby("analyst_nm")
             .agg(total_rev=("revenue_usd", "sum"), total_mtg=("meetings_cnt", "sum"))
             .reset_index())
    eff["rpm"] = (eff["total_rev"] / eff["total_mtg"].replace(0, 1)).round(0)
    eff = eff.sort_values("rpm", ascending=True)
    _bar(ws, "A59",
         eff["analyst_nm"].tolist(),
         [float(v) for v in eff["rpm"]],
         "Analyst Revenue per Meeting (USD) \u2014 Worst to Best",
         color=TEAL, width=CH_W_FULL, height=CH_H_FULL, horiz=True,
         show_labels=True, label_fmt='$#,##0')

    # ════════════════════════════════════════════════════════════════════════
    # MINI TABLES (cols Q-T, rows 32+)
    # ════════════════════════════════════════════════════════════════════════

    # KEY WATCHLIST header sits directly above the first mini table
    _merge(ws, 32, TBL_S, 32, TBL_E,
           "  KEY WATCHLIST",
           bg="0D1F3C", fnt=_font(10, True, GOLD), aln=_align(h="left"), h=22)

    # Table 1: Top 5 clients
    top5 = df.nlargest(5, "revenue_usd")[["client_name", "revenue_usd", "meetings_cnt"]]
    t1 = [(r.client_name[:20], f"${r.revenue_usd:,.0f}", int(r.meetings_cnt))
          for r in top5.itertuples()]
    next_row = _mini_table(ws, 33,
        ["Client", "Revenue", "Mtgs"], t1,
        "TOP 5 \u2014 REVENUE LEADERS", title_bg=NAVY)

    # Table 2: High effort / low revenue
    if "Engagement_Score" in df.columns:
        med_eng = df["Engagement_Score"].median()
        p25_rev = df["revenue_usd"].quantile(0.25)
        under = df[(df["Engagement_Score"] >= med_eng) & (df["revenue_usd"] <= p25_rev)].copy()
    else:
        med_mtg = df["meetings_cnt"].median()
        p25_rev = df["revenue_usd"].quantile(0.25)
        under = df[(df["meetings_cnt"] >= med_mtg) & (df["revenue_usd"] <= p25_rev)].copy()
    under["effort_1k"] = (under["meetings_cnt"] / (under["revenue_usd"] / 1000.0 + 0.01)).round(0)
    under = under.sort_values("effort_1k", ascending=False).head(7)
    t2 = [(r.client_name[:20], f"${r.revenue_usd:,.0f}", int(r.meetings_cnt), int(r.effort_1k))
          for r in under.itertuples()]
    next_row = _mini_table(ws, next_row,
        ["Client", "Revenue", "Mtgs", "Mtgs/$1K"], t2,
        "WASTED EFFORT \u2014 HIGH MTG, LOW REV", title_bg=AMBER,
        flag_style="amber", col_widths=tbl_widths)

    # Table 3: MiFID Expired
    exp_df = df[df["MiFID_Status"] == "EXPIRED"][
        ["client_name", "analyst_nm", "mifid_expiry_dt"]].head(7)
    t3 = [(r.client_name[:20], r.analyst_nm[:14], str(getattr(r, "mifid_expiry_dt", ""))[:10])
          for r in exp_df.itertuples()]
    next_row = _mini_table(ws, next_row,
        ["Client", "Analyst", "Expiry"], t3,
        "MIFID EXPIRED \u2014 ACTION REQUIRED", title_bg=RED,
        flag_style="red", col_widths=tbl_widths[:3])

    # Table 4: Bottom 5 efficiency
    rpm_col = "Rev_Per_Meeting" if "Rev_Per_Meeting" in df.columns else None
    if rpm_col:
        bot5 = df[df[rpm_col] > 0].nsmallest(5, rpm_col)[
            ["client_name", "revenue_usd", "meetings_cnt", rpm_col]]
        t4 = [(r.client_name[:20], f"${r.revenue_usd:,.0f}", int(r.meetings_cnt),
               round(getattr(r, rpm_col), 0)) for r in bot5.itertuples()]
    else:
        bot5 = df.copy()
        bot5["_rpm"] = bot5["revenue_usd"] / bot5["meetings_cnt"].replace(0, 1)
        bot5 = bot5[bot5["_rpm"] > 0].nsmallest(5, "_rpm")[
            ["client_name", "revenue_usd", "meetings_cnt", "_rpm"]]
        t4 = [(r.client_name[:20], f"${r.revenue_usd:,.0f}", int(r.meetings_cnt),
               f"${r._rpm:,.0f}") for r in bot5.itertuples()]
    _mini_table(ws, next_row,
        ["Client", "Revenue", "Mtgs", "Rev/Mtg"], t4,
        "LOWEST REVENUE EFFICIENCY", title_bg="2C4770",
        flag_style="red", col_widths=tbl_widths)

    # ── Footer ───────────────────────────────────────────────────────────────
    _merge(ws, 85, 1, 85, TBL_E,
           f"  Nomura CAMS Analytics Pipeline  \u00b7  {processing_date}"
           "  \u00b7  CONFIDENTIAL \u2014 For Authorised Recipients Only"
           "  \u00b7  Filter panel (top right) updates KPI metrics live.",
           bg="0D1F3C", fnt=_font(8, False, "A8C6E8"), aln=_align(h="left"), h=18)


# ─── Data Sheet ──────────────────────────────────────────────────────────────
def _build_data_sheet(ws, df, title_text):
    headers = list(df.columns)
    n_cols  = len(headers)

    # Row 1: title
    _merge(ws, 1, 1, 1, n_cols, title_text,
           bg=NAVY, fnt=_font(11, True, WHITE), aln=_align(), h=30)

    # Row 2: headers
    for ci, h in enumerate(headers, 1):
        _write(ws, 2, ci, h,
               bg=BLUE, fnt=_font(10, True, WHITE),
               aln=_align(), bdr=_thin(), h=20)

    # Data rows
    COL_MAP = {h: i for i, h in enumerate(headers, 1)}
    for ri, row_data in enumerate(df.itertuples(index=False), start=3):
        base_bg = WHITE if (ri % 2 == 1) else BLUE_LT
        mifid_val = str(getattr(row_data, "MiFID_Status", "")).upper()
        flag_val  = str(getattr(row_data, "Flag", ""))
        if mifid_val == "EXPIRED":
            row_bg = RED_LT; row_txt = RED_TXT; row_bold = True
        elif flag_val == "Review":
            row_bg = AMBER_LT; row_txt = AMBER_TXT; row_bold = True
        else:
            row_bg = base_bg; row_txt = "000000"; row_bold = False

        ws.row_dimensions[ri].height = 16
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = val if not (isinstance(val, float) and pd.isna(val)) else None
            c.fill      = _fill(row_bg)
            c.font      = _font(9, row_bold, row_txt)
            c.border    = _thin()
            c.alignment = _align(h="left", v="center")

            # Rev/Meeting low = extra highlight
            if headers[ci-1] == "Rev_Per_Meeting" and not row_bold:
                try:
                    rpm_v = float(val)
                    if 0 < rpm_v < 5000:
                        c.fill = _fill(RED_LT)
                        c.font = _font(9, True, RED_TXT)
                except (TypeError, ValueError):
                    pass

    # Column widths
    name_kws = ("name", "analyst", "tier", "region", "status", "flag")
    date_kws = ("date", "dt", "expiry")
    for ci, h in enumerate(headers, 1):
        hl = h.lower()
        if any(k in hl for k in name_kws):
            ws.column_dimensions[get_column_letter(ci)].width = 24
        elif any(k in hl for k in date_kws):
            ws.column_dimensions[get_column_letter(ci)].width = 14
        else:
            ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{len(df) + 2}"


# ─── Analyst Summary Sheet ───────────────────────────────────────────────────
def _build_analyst_sheet(ws, df):
    df = df[~df["analyst_nm"].isin(["System", "Unknown Analyst"])]
    agg = df.groupby("analyst_nm").agg(
        Clients   = ("client_name",   "count"),
        Revenue   = ("revenue_usd",   "sum"),
        Meetings  = ("meetings_cnt",  "sum"),
        Events    = ("events_attended", "sum") if "events_attended" in df.columns else ("meetings_cnt", "sum"),
        Flagged   = ("Flag", lambda x: (x == "Review").sum()),
        Expired   = ("MiFID_Status", lambda x: (x == "EXPIRED").sum()),
    ).reset_index()

    agg["Rev_Per_Meeting"] = (agg["Revenue"] / agg["Meetings"].replace(0, 1)).round(0)
    agg["Efficiency_Grade"] = agg["Rev_Per_Meeting"].apply(
        lambda v: "A — Excellent" if v >= 50000 else
                  "B — Good"      if v >= 20000 else
                  "C — Average"   if v >= 8000  else
                  "D — Poor"
    )
    agg = agg.sort_values("Rev_Per_Meeting", ascending=False).reset_index(drop=True)
    agg["Revenue"] = agg["Revenue"].round(0)

    headers = ["Analyst Name", "Clients", "Revenue (USD)", "Meetings",
               "Events", "Flagged", "MiFID Exp", "Rev/Meeting", "Grade"]
    col_data = ["analyst_nm", "Clients", "Revenue", "Meetings",
                "Events", "Flagged", "Expired", "Rev_Per_Meeting", "Efficiency_Grade"]

    _merge(ws, 1, 1, 1, len(headers), "ANALYST PERFORMANCE SUMMARY",
           bg=NAVY, fnt=_font(11, True, WHITE), aln=_align(), h=30)
    for ci, h in enumerate(headers, 1):
        _write(ws, 2, ci, h,
               bg=BLUE, fnt=_font(10, True, WHITE),
               aln=_align(), bdr=_thin(), h=20)

    for ri, row in enumerate(agg.itertuples(), start=3):
        grade = getattr(row, "Efficiency_Grade", "")
        if "A" in grade:
            bg_r = GREEN_LT; txt_r = GREEN_TXT
        elif "D" in grade:
            bg_r = RED_LT;   txt_r = RED_TXT
        else:
            bg_r = WHITE if ri % 2 == 1 else BLUE_LT
            txt_r = "000000"
        ws.row_dimensions[ri].height = 16
        for ci, col in enumerate(col_data, 1):
            val = getattr(row, col, "")
            c = ws.cell(row=ri, column=ci)
            c.value = (f"${val:,.0f}" if col in ("Revenue", "Rev_Per_Meeting") and
                       isinstance(val, (int, float)) else val)
            c.fill      = _fill(bg_r)
            c.font      = _font(9, False, txt_r)
            c.border    = _thin()
            c.alignment = _align(h="left" if ci == 1 else "right")

    col_widths = [28, 9, 16, 10, 9, 9, 10, 16, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"


def _build_jdf_alignment_sheet(ws, df, metadata):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    widths = [22, 18, 18, 44, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    _banner(ws, 1, 1, 5,
            "OVERVIEW  ·  HOW THE WORKBOOK MAPS TO THE ROLE BRIEF",
            NAVY_DK, WHITE, size=14, bold=True, height=34)
    _banner(ws, 2, 1, 5,
            "This sheet makes the Fixed Income client management use case explicit: performance, access control, focus lists, and engagement analytics.",
            BLUE_LT, SLATE, size=9, italic=True, bold=False, height=22)

    headers = ["JDF Theme", "Metric", "Value", "Why it matters", "Source"]
    for ci, h in enumerate(headers, 1):
        _write(ws, 4, ci, h, bg=NAVY, fnt=_font(9, True, WHITE), aln=_align(), bdr=_thin(), h=19)

    total_clients = int(metadata.get("total_rows_clean", len(df)))
    total_revenue = float(df["revenue_usd"].sum())
    expired_mifid = int(metadata.get("expired_mifid_count", int((df["MiFID_Status"] == "EXPIRED").sum())))
    focus_list_count = int(
        metadata.get(
            "focus_list_count",
            int((df["Client_Focus_Category"] == "Focus List").sum() if "Client_Focus_Category" in df.columns else (df["Flag"] == "Review").sum()),
        )
    )
    access_granted = int(
        metadata.get(
            "research_access_granted_count",
            int((df["Research_Access_Status"] == "Granted").sum() if "Research_Access_Status" in df.columns else (df["MiFID_Status"] != "EXPIRED").sum()),
        )
    )
    access_revoked = int(
        metadata.get(
            "research_access_revoked_count",
            int((df["Research_Access_Status"] == "Revoked").sum() if "Research_Access_Status" in df.columns else (df["MiFID_Status"] == "EXPIRED").sum()),
        )
    )
    avg_readership = float(
        metadata.get(
            "avg_readership_rate",
            round(float(df["Readership_Rate"].mean()), 2) if "Readership_Rate" in df.columns and len(df)
            else round(float((df["reports_read"] / df["meetings_cnt"].replace(0, 1)).mean()), 2) if "reports_read" in df.columns and "meetings_cnt" in df.columns and len(df)
            else 0.0,
        )
    )
    top_analyst = metadata.get("top_analyst_name", "")
    if not top_analyst and "analyst_nm" in df.columns and "revenue_usd" in df.columns:
        grouped = df.groupby("analyst_nm")["revenue_usd"].sum()
        top_analyst = grouped.idxmax() if not grouped.empty else "N/A"
    top_analyst = top_analyst or "N/A"

    worst_analyst = metadata.get("worst_analyst_name", "")
    if not worst_analyst and "analyst_nm" in df.columns and "Rev_Per_Meeting" in df.columns:
        non_zero = df[df["Rev_Per_Meeting"] > 0]
        if not non_zero.empty:
            grouped = non_zero.groupby("analyst_nm")["Rev_Per_Meeting"].mean()
            worst_analyst = grouped.idxmin() if not grouped.empty else "N/A"
    worst_analyst = worst_analyst or "N/A"

    rows = [
        ("Portfolio coverage", "Client universe", total_clients, "Measures the size of the coverage universe analysed.", "Data / Dashboard"),
        ("Commercial outcome", "Total revenue", f"${total_revenue:,.0f}", "Shows monetisation across the client base.", "Data / Dashboard"),
        ("Performance analysis", "Top analyst", top_analyst, "Identifies the strongest revenue contributor.", "Analyst_Summary"),
        ("Performance analysis", "Lowest efficiency analyst", worst_analyst, "Flags the weakest revenue per meeting performance.", "Analyst_Summary"),
        ("Compliance", "Expired MiFID", expired_mifid, "Highlights clients needing access review.", "Data / Flagged_Clients"),
        ("Access tracking", "Research access granted", access_granted, "Shows clients currently allowed to access research.", "Data"),
        ("Access tracking", "Research access revoked", access_revoked, "Shows where access has been revoked after expiry.", "Data"),
        ("Client focus", "Focus list clients", focus_list_count, "Surfaces accounts requiring senior review.", "Flagged_Clients"),
        ("Engagement", "Average readership rate", f"{avg_readership:.2f}", "Maps content consumption against effort and revenue.", "Data / Efficiency_Analysis"),
    ]

    for idx, (theme, metric, value, why, source) in enumerate(rows, start=5):
        bg = WHITE if idx % 2 == 1 else BLUE_LT
        _write(ws, idx, 1, theme, bg=bg, fnt=_font(9, True, NAVY), aln=_align(h="left"), bdr=_thin(), h=18)
        _write(ws, idx, 2, metric, bg=bg, fnt=_font(9, False, SLATE), aln=_align(h="left"), bdr=_thin(), h=18)
        _write(ws, idx, 3, value, bg=bg, fnt=_font(9, True, NAVY_DK), aln=_align(), bdr=_thin(), h=18)
        _write(ws, idx, 4, why, bg=bg, fnt=_font(9, False, SLATE), aln=_align(h="left", wrap=True), bdr=_thin(), h=18)
        _write(ws, idx, 5, source, bg=bg, fnt=_font(9, False, AMBER), aln=_align(), bdr=_thin(), h=18)

    _merge(ws, 16, 1, 16, 5,
           "Workbook schema note: Data now includes Research_Access_Status, Client_Focus_Category, and Readership_Rate to align more closely to the JDF.",
           bg=BLUE_PAL, fnt=_font(8.5, False, SLATE, italic=True), aln=_align(), h=18)

    ws.freeze_panes = "A5"


# ─── Flagged Clients Sheet ───────────────────────────────────────────────────
def _build_flagged_sheet(ws, df):
    flagged = df[df["Flag"] == "Review"].reset_index(drop=True)
    headers = list(flagged.columns)
    n_cols  = len(headers)

    _merge(ws, 1, 1, 1, n_cols,
           "CLIENTS REQUIRING SENIOR MANAGEMENT REVIEW — HIGH MEETINGS, LOW REVENUE",
           bg=RED, fnt=_font(11, True, WHITE), aln=_align(), h=30)
    for ci, h in enumerate(headers, 1):
        _write(ws, 2, ci, h,
               bg=BLUE, fnt=_font(10, True, WHITE),
               aln=_align(), bdr=_thin(), h=20)

    for ri, row in enumerate(flagged.itertuples(index=False), start=3):
        ws.row_dimensions[ri].height = 16
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci)
            c.value     = val if not (isinstance(val, float) and pd.isna(val)) else None
            c.fill      = _fill(AMBER_LT)
            c.font      = _font(9, True, AMBER_TXT)
            c.border    = _thin()
            c.alignment = _align(h="left", v="center")

    name_kws = ("name", "analyst", "tier", "region", "status", "flag")
    for ci, h in enumerate(headers, 1):
        w = 24 if any(k in h.lower() for k in name_kws) else 14
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"


# ─── Efficiency / ROI Analysis Sheet ─────────────────────────────────────────
def _build_efficiency_sheet(ws, df):
    """
    Shows every client ranked by Revenue/Engagement efficiency.
    Highlights low-ROI clients (high effort, low revenue).
    """
    ef = df.copy()

    if "Engagement_Score" in ef.columns:
        ef["Eng_Score"]    = ef["Engagement_Score"].round(1)
        ef["Rev_Per_Eng"]  = (ef["revenue_usd"] / ef["Engagement_Score"].replace(0, 0.01)).round(0)
    else:
        ef["Eng_Score"]    = ef["meetings_cnt"]
        ef["Rev_Per_Eng"]  = (ef["revenue_usd"] / ef["meetings_cnt"].replace(0, 1)).round(0)

    ef["Meetings_per_1k_Rev"] = (
        ef["meetings_cnt"] / (ef["revenue_usd"] / 1000.0 + 0.01)
    ).round(1)

    ef["ROI_Rating"] = ef["Rev_Per_Eng"].apply(
        lambda v: "Excellent" if v >= 15000 else
                  "Good"      if v >= 5000  else
                  "Average"   if v >= 1000  else
                  "Poor — Review"
    )

    ef = ef.sort_values("Rev_Per_Eng", ascending=True).reset_index(drop=True)

    out_cols  = ["client_name", "analyst_nm", "REGION", "client_tier",
                 "revenue_usd", "meetings_cnt", "Eng_Score",
                 "Rev_Per_Meeting", "Rev_Per_Eng",
                 "Meetings_per_1k_Rev", "MiFID_Status", "Flag", "ROI_Rating"]
    out_cols  = [c for c in out_cols if c in ef.columns]
    ef_out    = ef[out_cols].copy()

    display_headers = {
        "client_name":         "Client Name",
        "analyst_nm":          "Analyst",
        "REGION":              "Region",
        "client_tier":         "Tier",
        "revenue_usd":         "Revenue (USD)",
        "meetings_cnt":        "Meetings",
        "Eng_Score":           "Engagement Score",
        "Rev_Per_Meeting":     "Rev / Meeting",
        "Rev_Per_Eng":         "Rev / Eng Point",
        "Meetings_per_1k_Rev": "Meetings per $1K Rev",
        "MiFID_Status":        "MiFID",
        "Flag":                "Flag",
        "ROI_Rating":          "ROI Rating",
    }
    headers = [display_headers.get(c, c) for c in out_cols]
    n_cols  = len(headers)

    # Summary stats row at top
    _merge(ws, 1, 1, 1, n_cols,
           "CLIENT ENGAGEMENT ROI & EFFICIENCY ANALYSIS — Sorted: Worst → Best",
           bg=NAVY, fnt=_font(12, True, WHITE), aln=_align(), h=32)

    # Stats bar
    p25 = ef["Rev_Per_Eng"].quantile(0.25)
    med = ef["Rev_Per_Eng"].median()
    p75 = ef["Rev_Per_Eng"].quantile(0.75)
    _merge(ws, 2, 1, 2, n_cols,
           f"Portfolio Median Rev/Eng: ${med:,.0f}  |  "
           f"Bottom 25%: ${p25:,.0f}  |  Top 25%: ${p75:,.0f}  |  "
           f"Clients Below Median: {int((ef['Rev_Per_Eng'] < med).sum())}  |  "
           f"Clients Flagged (Poor ROI): {int((ef['ROI_Rating'] == 'Poor — Review').sum())}",
           bg=BLUE_PAL, fnt=_font(9, False, NAVY, italic=True), aln=_align(), h=20)

    # Headers
    for ci, h in enumerate(headers, 1):
        _write(ws, 3, ci, h,
               bg=BLUE, fnt=_font(10, True, WHITE),
               aln=_align(), bdr=_thin(), h=20)

    # Data
    for ri, row in enumerate(ef_out.itertuples(index=False), start=4):
        rating = str(getattr(row, "ROI_Rating", "")) if "ROI_Rating" in out_cols else ""
        if "Poor" in rating:
            bg_r, txt_r, bld = RED_LT, RED_TXT, True
        elif rating == "Excellent":
            bg_r, txt_r, bld = GREEN_LT, GREEN_TXT, False
        elif rating == "Good":
            bg_r, txt_r, bld = BLUE_LT, NAVY, False
        else:
            bg_r, txt_r, bld = WHITE if ri % 2 == 0 else GREY_LT, "000000", False

        ws.row_dimensions[ri].height = 15
        for ci, (col, val) in enumerate(zip(out_cols, row), 1):
            c = ws.cell(row=ri, column=ci)
            if isinstance(val, float) and pd.isna(val):
                c.value = None
            elif col in ("revenue_usd", "Rev_Per_Meeting", "Rev_Per_Eng"):
                c.value = f"${float(val):,.0f}" if val else "$0"
            else:
                c.value = val
            c.fill      = _fill(bg_r)
            c.font      = _font(9, bld, txt_r)
            c.border    = _thin()
            c.alignment = _align(h="left" if ci <= 4 else "right")

    # Column widths
    w_map = {
        "client_name": 28, "analyst_nm": 22, "REGION": 12, "client_tier": 14,
        "revenue_usd": 16, "meetings_cnt": 12, "Eng_Score": 16,
        "Rev_Per_Meeting": 16, "Rev_Per_Eng": 16, "Meetings_per_1k_Rev": 22,
        "MiFID_Status": 14, "Flag": 12, "ROI_Rating": 18,
    }
    for ci, col in enumerate(out_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w_map.get(col, 14)

    ws.freeze_panes = "A4"


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        raise ValueError("Usage: python formatting.py <clean_file_path> '<json_metadata>'")

    clean_file_path = sys.argv[1]
    metadata        = json.loads(sys.argv[2])
    total_rows_clean = metadata.get("total_rows_clean", "N/A")
    processing_date  = metadata.get("processing_date", datetime.now().strftime("%d %B %Y"))
    input_timestamp  = metadata.get("timestamp", datetime.now().strftime("%Y%m%d_%H%M%S"))

    output_path = f"/Users/priyanshupatel/.n8n-files/Nomura_Formatted_{input_timestamp}.xlsx"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Load clean data (preprocessing saves sheet as "Data")
    df = pd.read_excel(clean_file_path, engine="openpyxl", sheet_name=0)

    title_text = (
        f"NOMURA FI CLIENT ENGAGEMENT REPORT  |  "
        f"Processed: {processing_date}  |  {total_rows_clean} records"
    )

    wb = load_workbook(clean_file_path)

    # ── Sheet: Data ──────────────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Data"
    _build_data_sheet(ws_data, df, title_text)

    # Hidden helper lists for the interactive filter sheet
    ws_lists = wb.create_sheet("Filter_Lists")
    list_ranges = _build_filter_lists_sheet(ws_lists, df)

    # ── Sheet: Analyst_Summary ───────────────────────────────────────────────
    for name in ["Analyst_Summary", "Efficiency_Analysis", "Flagged_Clients", "Overview", "Dashboard"]:
        if name in wb.sheetnames:
            del wb[name]

    ws_analyst = wb.create_sheet("Analyst_Summary")
    _build_analyst_sheet(ws_analyst, df)

    # ── Sheet: Overview ─────────────────────────────────────────────────────
    ws_jdf = wb.create_sheet("Overview")
    _build_jdf_alignment_sheet(ws_jdf, df, metadata)

    # ── Sheet: Flagged_Clients ───────────────────────────────────────────────
    ws_flagged = wb.create_sheet("Flagged_Clients")
    _build_flagged_sheet(ws_flagged, df)

    # ── Sheet: Efficiency_Analysis ───────────────────────────────────────────
    ws_eff = wb.create_sheet("Efficiency_Analysis")
    _build_efficiency_sheet(ws_eff, df)

    # ── Sheet: _ChartData (hidden, stores all chart source data) ────────────
    ws_cd = wb.create_sheet("_ChartData")
    ws_cd.sheet_state = "hidden"
    ws_cd.column_dimensions["A"].width = 30
    ws_cd.column_dimensions["B"].width = 18
    _chart_ws[0]  = ws_cd
    _chart_off[0] = 0

    # ── Sheet: Dashboard (inserted first) ────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard", 0)
    _build_dashboard(ws_dash, df, processing_date, total_rows_clean, list_ranges)
    wb.active = ws_dash
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_path)

    print(json.dumps({
        "formatted_file_path": output_path,
        "success": True,
        "total_rows_formatted": len(df),
        "sheets": ["Dashboard", "Data", "Analyst_Summary", "Overview", "Flagged_Clients", "Efficiency_Analysis"],
        "output_timestamp": input_timestamp,
    }))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        tb = traceback.format_exc()
        logging.error("Formatting failed:\n%s", tb)
        print(json.dumps({
            "formatted_file_path": None,
            "success": False,
            "error": str(exc),
            "error_detail": tb,
        }))
        sys.exit(1)
